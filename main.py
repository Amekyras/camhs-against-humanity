import openpyxl
import random
import os
from gtts import gTTS
from playsound import playsound
import glob
import time


def clearscreen():
    os.system('cls' if os.name == 'nt' else 'clear')


def prep_black(black):
    prepb = black.text.replace("______", "BLANK")
    tts = gTTS(prepb, lang="en", tld="co.uk")
    tts.save("black.mp3")


def prep_white(black, whites):
    prepw = ""
    prepb = black.text.replace("______", "BLANK")
    if "BLANK" not in prepb:
        prepb = prepb + " BLANK"
    for j in range(0, black.pc):
        prepw = prepb.replace("BLANK", whites.play[j], 1)
    tts = gTTS(prepw, lang="en", tld="co.uk")
    tts.save(f"{str(whites.pb.name)}.mp3")


class Card:
    def __init__(self, card_colour, card_id, card_options, card_text):
        self.colour = card_colour
        self.id = card_id
        self.pc = int(card_options)
        self.text = card_text


def prep_winner(name):
    tts = gTTS(f"{name} won this round with.", lang="en", tld="co.uk")
    tts.save(f"{name}win.mp3")


class Player:
    def __init__(self, player_id, player_is_human, player_score, player_is_czar, player_name):
        self.id = player_id
        self.human = player_is_human
        self.score = player_score
        self.czar = player_is_czar
        self.name = player_name
        self.hand = []
        if speaker:
            prep_winner(player_name)

    def add_score(self):
        self.score += 1

    def reset_score(self):
        self.score = 0

    def add_card(self):
        self.hand.append(white_cards.pop())


class Entry:
    def __init__(self, played_by):
        self.pb = played_by
        self.play = []

    def enter(self, entry):
        card_entry = active_player.hand.pop(entry)
        self.play.append(card_entry.text)


# Card loader and settings
white_card_count = 231
black_card_count = 116
white_card_column = 1
black_card_column = 2
option_column = 3
start_row = 2
white_cards = []
black_cards = []
white_cards_fresh = []
black_cards_fresh = []

card_lib = openpyxl.load_workbook("data.xlsx")
card_sheet = card_lib.active

for i in range(0, white_card_count):
    white_cards_fresh.append(
        Card("White", str(i) + "W", 0, card_sheet.cell(row=(i + start_row), column=white_card_column).value))

for i in range(0, black_card_count):
    black_cards_fresh.append(Card("Black", str(i) + "B", card_sheet.cell(row=i + start_row, column=option_column).value,
                                  card_sheet.cell(row=i + start_row, column=black_card_column).value))

# Cleanup MP3s from last run
globs = glob.glob("*.mp3")
while len(globs) > 0:
    os.remove(globs[0])
    globs = glob.glob("*.mp3")


# Start game
print("Welcome to CAMHS Against Humanity! "
      "\nCards written by the inmates of Cotswold Spa Hospital and associates"
      "\nCoded by amekyras"
      "\nhttps://github.com/Amekyras/camhs-against-humanity")

# Game setup
player_num = 3
hand_size = 10
rounds = 0
while True:
    try:
        player_num = int(input("How many people are playing?"))
    except (TypeError, ValueError):
        continue
    if player_num < 3:
        continue
    break

while True:
    try:
        hand_size = int(input("How many cards in a hand?"))
    except (TypeError, ValueError):
        continue
    break

speaker_prompt = ""
while True:
    try:
        speaker_prompt = input("Would you like to enable TTS?")
    except (TypeError, ValueError):
        continue
    break

affirm = ["", "Yes", "yes", "Y", "y"]
if speaker_prompt in affirm:
    speaker = True
else:
    speaker = False
white_cards = white_cards_fresh.copy()
black_cards = black_cards_fresh.copy()

# Player setup
players = []
for i in range(0, player_num):
    print("Player", i)
    players.append(Player(i, True, 0, False, input("What is your name?")))
czar_counter = 0
clearscreen()

input("Press enter to begin the game.")
# Main game loop
while True:
    # Check if any players have empty hands

    # Give players cards
    if len(white_cards) > 0:
        random.shuffle(white_cards)
        for i in range(0, len(players)):
            while len(players[i].hand) < hand_size:
                players[i].add_card()

    # remove czar
    for i in range(0, player_num):
        players[i].czar = False

    # Round Counter
    rounds += 1
    clearscreen()
    print(f"Welcome to Round {rounds}")

    # add czar
    czar_pick = players[czar_counter]
    if czar_counter >= (len(players)-1):
        czar_counter = 0
    else:
        czar_counter += 1
    czar_pick.czar = True
    print(f"{czar_pick.name} is the Card Czar this round!")

    # pick black card
    black_pick = black_cards.pop(random.randint(0, len(black_cards)))
    if speaker:
        prep_black(black_pick)
    print(f"This round's prompt is\n"
          f"{black_pick.text}"
          f"\nThis card has {black_pick.pc} prompt(s)")

    # Round
    entries = []
    for i in range(0, player_num):
        active_player = players[i]
        if active_player.czar is False:
            entry_count = 0
            # Play loop
            print(f"{active_player.name}, it's your turn! Press enter to continue.")
            input()
            clearscreen()
            while entry_count < black_pick.pc:
                print(black_pick.text)
                for x in range(0, len(active_player.hand)):
                    print(f"Card {x} - {active_player.hand[x].text}")
                print(f"Please select a card. This is Entry {entry_count + 1} of {black_pick.pc}")
                pick = 0
                while True:
                    try:
                        pick = int(input())
                        if pick > (len(active_player.hand)-1):
                            print("Invalid input")
                            continue
                    except (TypeError, ValueError):
                        continue
                    break
                entry_count += 1
                if entry_count == 1:
                    entries.append(Entry(active_player))
                entries[-1].enter(pick)
                clearscreen()
            if speaker:
                prep_white(black_pick, entries[-1])

    random.shuffle(entries)
    clearscreen()
    print(f"{czar_pick.name}, it's time to pick the winner!")
    print(black_pick.text)
    for i in range(0, len(entries)):
        print(f"Entry {i} - " + ", ".join(entries[i].play))
    # GENERATE TTS HERE
    if speaker:
        for i in range(0, len(entries)):
            playsound(f"{entries[i].pb.name}.mp3")
            time.sleep(0.1)
    winner = 0
    while True:
        try:
            winner = int(input())
        except(TypeError, ValueError):
            continue
        break
    winning_card = entries[winner].play
    winning_player = entries[winner].pb
    if speaker:
        playsound(f"{winning_player.name}win.mp3")
        playsound(f"{winning_player.name}.mp3")
    print(winning_player.name, "won this round with: " + ", ".join(winning_card))
    winning_player.add_score()
    print("Scores:")
    for i in range(0, len(players)):
        print(f"{players[i].name} :  {players[i].score}")
    for i in range(0, len(players)):
        os.remove(f"{players[i].name}.mp3")
    input("Press enter to begin the next round.")
